#!/usr/bin/env python3
"""
MICROHUB EXCEL EXPORTER v3.0
============================
Export papers to Excel (.xlsx) format for WordPress import.
Updated for v3 scraper with full text, figures, and supplementary materials.

Usage:
  python microhub_export_excel_v3.py                     # Export all papers
  python microhub_export_excel_v3.py --limit 10000       # Limit papers
  python microhub_export_excel_v3.py --full-text-only    # Only papers with full text
  python microhub_export_excel_v3.py --with-figures      # Only papers with figures
"""

import sqlite3
import json
import argparse
import logging
import os
from typing import List, Dict, Any
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl is required for Excel export.")
    print("Install it with: pip install openpyxl")
    exit(1)

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(message)s')
logger = logging.getLogger(__name__)


class ExcelExporterV3:
    """Export papers to Excel format for WordPress - v3 with full text support."""

    def __init__(self, db_path: str = 'microhub.db'):
        self.db_path = db_path

    def get_json_field(self, row: Dict, field_name: str) -> List:
        """Safely parse JSON field."""
        try:
            value = row.get(field_name)
            if value:
                return json.loads(value)
            return []
        except (TypeError, json.JSONDecodeError):
            return []

    def get_field(self, row: Dict, field_name: str, default: Any = '') -> Any:
        """Safely get field value."""
        return row.get(field_name) or default

    def format_list_for_wordpress(self, items: List) -> str:
        """Format list as pipe-separated string for WordPress taxonomy import."""
        if not items:
            return ''
        if isinstance(items[0], dict):
            names = []
            for item in items:
                if 'source' in item:
                    names.append(item['source'])
                elif 'name' in item:
                    names.append(item['name'])
                elif 'id' in item:
                    names.append(str(item['id']))
                elif 'label' in item:
                    names.append(item['label'])
            return '|'.join(names)
        return '|'.join(str(item) for item in items)

    def format_urls(self, items: List) -> str:
        """Extract URLs from list of dicts."""
        if not items:
            return ''
        urls = []
        for item in items:
            if isinstance(item, dict):
                url = item.get('url') or item.get('image_url')
                if url:
                    urls.append(url)
        return '|'.join(urls)

    def format_figures_for_display(self, figures: List) -> str:
        """Format figures with captions for display."""
        if not figures:
            return ''
        formatted = []
        for fig in figures:
            if isinstance(fig, dict):
                label = fig.get('label', '')
                title = fig.get('title', '')
                caption = fig.get('caption', '')[:200] if fig.get('caption') else ''  # Truncate caption
                parts = [p for p in [label, title, caption] if p]
                if parts:
                    formatted.append(' - '.join(parts))
        return '\n'.join(formatted[:10])  # Max 10 figures for display

    def export(self, output_path: str = 'microhub_papers_v3.xlsx',
               limit: int = None,
               enriched_only: bool = False,
               full_text_only: bool = False,
               with_figures_only: bool = False,
               chunk_size: int = 50000):
        """Export papers to Excel in chunks."""

        logger.info("=" * 60)
        logger.info("MICROHUB EXCEL EXPORTER v3.0 - FULL TEXT EDITION")
        logger.info("=" * 60)

        conn = sqlite3.connect(self.db_path, timeout=120.0)
        conn.row_factory = sqlite3.Row

        # Build query conditions
        conditions = []
        if enriched_only:
            conditions.append("enriched_at IS NOT NULL")
        if full_text_only:
            conditions.append("has_full_text = 1")
        if with_figures_only:
            conditions.append("has_figures = 1")

        where_clause = " AND ".join(conditions) if conditions else "1=1"

        query = f"""
            SELECT * FROM papers
            WHERE {where_clause}
            ORDER BY priority_score DESC, citation_count DESC, year DESC
        """

        if limit:
            query += f" LIMIT {limit}"

        logger.info("Fetching papers from database...")
        cursor = conn.execute(query)

        chunk_num = 1
        chunk_papers = 0
        papers_written = 0
        created_files = []

        base_name = output_path.replace('.xlsx', '')

        def create_new_workbook():
            wb = Workbook()

            # Main papers sheet
            ws_main = wb.active
            ws_main.title = "Papers"

            main_columns = [
                ('post_title', 'Title', 80),
                ('post_content', 'Abstract', 100),
                ('methods', 'Methods', 80),
                ('post_status', 'Status', 10),
                ('post_type', 'Type', 15),
                ('pmid', 'PMID', 12),
                ('doi', 'DOI', 30),
                ('pmc_id', 'PMC ID', 15),
                ('authors', 'Authors', 50),
                ('journal', 'Journal', 40),
                ('year', 'Year', 8),
                ('doi_url', 'DOI URL', 40),
                ('pubmed_url', 'PubMed URL', 45),
                ('pmc_url', 'PMC URL', 50),
                ('github_url', 'GitHub URL', 50),
                ('citation_count', 'Citations', 10),
                ('priority_score', 'Priority', 10),
                ('microscopy_techniques', 'Microscopy Techniques', 40),
                ('microscope_brands', 'Microscope Brands', 30),
                ('microscope_models', 'Microscope Models', 40),
                ('image_analysis_software', 'Image Analysis Software', 40),
                ('image_acquisition_software', 'Image Acquisition Software', 40),
                ('sample_preparation', 'Sample Preparation', 40),
                ('fluorophores', 'Fluorophores', 40),
                ('organisms', 'Organisms', 30),
                ('protocols', 'Protocols', 50),
                ('protocols_urls', 'Protocol URLs', 60),
                ('repositories', 'Data Repositories', 40),
                ('repositories_urls', 'Repository URLs', 60),
                ('supplementary', 'Supplementary Materials', 50),
                ('rrids', 'RRIDs', 40),
                ('rrids_urls', 'RRID URLs', 60),
                ('rors', 'RORs', 30),
                ('figure_count', 'Figure Count', 12),
                ('figure_urls', 'Figure URLs', 80),
                ('figure_captions', 'Figure Info', 100),
                ('has_full_text', 'Has Full Text', 12),
                ('has_figures', 'Has Figures', 12),
                ('has_protocols', 'Has Protocols', 12),
                ('has_github', 'Has GitHub', 12),
                ('has_data', 'Has Data', 10),
            ]

            # Style header
            header_font = Font(bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='2E7D32', end_color='2E7D32', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            for col_idx, (field, header, width) in enumerate(main_columns, 1):
                cell = ws_main.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                ws_main.column_dimensions[get_column_letter(col_idx)].width = width

            ws_main.freeze_panes = 'A2'

            return wb, ws_main, main_columns

        wb, ws_main, main_columns = create_new_workbook()
        row_num = 2

        for row in cursor:
            row_dict = dict(row)

            # Parse JSON fields
            microscopy_techniques = self.get_json_field(row_dict, 'microscopy_techniques')
            microscope_brands = self.get_json_field(row_dict, 'microscope_brands')
            microscope_models = self.get_json_field(row_dict, 'microscope_models')
            image_analysis_software = self.get_json_field(row_dict, 'image_analysis_software')
            image_acquisition_software = self.get_json_field(row_dict, 'image_acquisition_software')
            sample_preparation = self.get_json_field(row_dict, 'sample_preparation')
            fluorophores = self.get_json_field(row_dict, 'fluorophores')
            organisms = self.get_json_field(row_dict, 'organisms')
            protocols = self.get_json_field(row_dict, 'protocols')
            repositories = self.get_json_field(row_dict, 'repositories')
            supplementary = self.get_json_field(row_dict, 'supplementary_materials')
            rrids = self.get_json_field(row_dict, 'rrids')
            rors = self.get_json_field(row_dict, 'rors')
            figures = self.get_json_field(row_dict, 'figures')

            # Build row data
            row_data = [
                self.get_field(row_dict, 'title'),
                self.get_field(row_dict, 'abstract'),
                self.get_field(row_dict, 'methods'),
                'publish',
                'microhub_paper',
                self.get_field(row_dict, 'pmid'),
                self.get_field(row_dict, 'doi'),
                self.get_field(row_dict, 'pmc_id'),
                self.get_field(row_dict, 'authors'),
                self.get_field(row_dict, 'journal'),
                self.get_field(row_dict, 'year'),
                self.get_field(row_dict, 'doi_url'),
                self.get_field(row_dict, 'pubmed_url'),
                self.get_field(row_dict, 'pmc_url'),
                self.get_field(row_dict, 'github_url'),
                self.get_field(row_dict, 'citation_count', 0),
                self.get_field(row_dict, 'priority_score', 0),
                self.format_list_for_wordpress(microscopy_techniques),
                self.format_list_for_wordpress(microscope_brands),
                self.format_list_for_wordpress(microscope_models),
                self.format_list_for_wordpress(image_analysis_software),
                self.format_list_for_wordpress(image_acquisition_software),
                self.format_list_for_wordpress(sample_preparation),
                self.format_list_for_wordpress(fluorophores),
                self.format_list_for_wordpress(organisms),
                self.format_list_for_wordpress(protocols),
                self.format_urls(protocols),
                self.format_list_for_wordpress(repositories),
                self.format_urls(repositories),
                self.format_list_for_wordpress(supplementary),
                self.format_list_for_wordpress(rrids),
                self.format_urls(rrids),
                self.format_list_for_wordpress(rors),
                self.get_field(row_dict, 'figure_count', 0),
                self.format_urls(figures),
                self.format_figures_for_display(figures),
                1 if self.get_field(row_dict, 'has_full_text') else 0,
                1 if figures else 0,
                1 if protocols else 0,
                1 if self.get_field(row_dict, 'github_url') else 0,
                1 if repositories else 0,
            ]

            for col_idx, value in enumerate(row_data, 1):
                cell = ws_main.cell(row=row_num, column=col_idx, value=value)
                cell.alignment = Alignment(vertical='top', wrap_text=True)

            row_num += 1
            papers_written += 1
            chunk_papers += 1

            if papers_written % 1000 == 0:
                logger.info(f"Processed {papers_written:,} papers...")

            if chunk_papers >= chunk_size:
                chunk_filename = f"{base_name}_chunk_{chunk_num}.xlsx"
                wb.save(chunk_filename)
                created_files.append(chunk_filename)
                logger.info(f"Saved chunk {chunk_num}: {chunk_filename} ({chunk_papers:,} papers)")

                chunk_num += 1
                chunk_papers = 0
                wb, ws_main, main_columns = create_new_workbook()
                row_num = 2

        # Save final chunk
        if chunk_papers > 0:
            chunk_filename = f"{base_name}_chunk_{chunk_num}.xlsx"
            wb.save(chunk_filename)
            created_files.append(chunk_filename)
            logger.info(f"Saved chunk {chunk_num}: {chunk_filename} ({chunk_papers:,} papers)")

        conn.close()

        total_size = sum(os.path.getsize(f) for f in created_files) / 1024 / 1024

        logger.info("=" * 60)
        logger.info("EXCEL EXPORT COMPLETE")
        logger.info("=" * 60)
        logger.info(f"Total papers exported: {papers_written:,}")
        logger.info(f"Number of chunk files: {len(created_files)}")
        logger.info(f"Total size: {total_size:.1f} MB")
        logger.info("")
        logger.info("Created files:")
        for f in created_files:
            size_mb = os.path.getsize(f) / 1024 / 1024
            logger.info(f"  {f} ({size_mb:.1f} MB)")

        return papers_written


def main():
    parser = argparse.ArgumentParser(description='Export papers to Excel v3.0')
    parser.add_argument('--db', default='microhub.db', help='Database path')
    parser.add_argument('--output', '-o', default='microhub_papers_v3.xlsx', help='Base output filename')
    parser.add_argument('--limit', type=int, help='Limit number of papers')
    parser.add_argument('--enriched-only', action='store_true', help='Only export enriched papers')
    parser.add_argument('--full-text-only', action='store_true', help='Only export papers with full text')
    parser.add_argument('--with-figures', action='store_true', help='Only export papers with figures')
    parser.add_argument('--chunk-size', type=int, default=50000, help='Papers per chunk file')

    args = parser.parse_args()

    exporter = ExcelExporterV3(db_path=args.db)
    exporter.export(
        output_path=args.output,
        limit=args.limit,
        enriched_only=args.enriched_only,
        full_text_only=args.full_text_only,
        with_figures_only=args.with_figures,
        chunk_size=args.chunk_size
    )


if __name__ == '__main__':
    main()